#导入excel操作库
from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelHelper(object):
    def __init__(self,file,head):
        self.fileStream = Workbook()
        #head = ['编号', '电流（mA）', '最小功率（W）', '最大功率（W）', '平均功率（W）', 'Std.Dev', '长度', '测试日期', '测试时长65秒，前10秒不取值']
        self.file = file
        if(head != None):
            writer = self.fileStream.active
            writer.append(head)
            self.fileStream.save(file)
    def writeline(self,rowData):
        #需要重新加载文件，然后再写入
        self.fileStream = load_workbook(self.file)
        writer = self.fileStream.active
        writer.append(rowData)
        self.fileStream.save(self.file)
    #读取表格中的所有数据
    def readall(self):
        wb = load_workbook(self.file)
        sheet = wb['Sheet']
        test_data = []  # 数据存储为列表格式
        for i in range(1, sheet.max_row):
            sub_data = []
            for j in range(1, 20):
                sub_data.append(sheet.cell(i+1,j).value)
            test_data.append(sub_data)
        return test_data  # 返回获取到的数据
        pass

